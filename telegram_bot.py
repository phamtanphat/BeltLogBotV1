import telebot
from telebot import types
import os
import generate_report
import history_manager
from datetime import datetime
from keep_alive import keep_alive

# ─── Cấu hình ───────────────────────────────────────────
TOKEN = os.environ.get("BOT_TOKEN", "8730792867:AAFyKUfJvnoF0uIZM5xGV3O9JtJcq7WdbUc")
bot = telebot.TeleBot(TOKEN)

DOWNLOAD_DIR = "downloads"
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# Dọn rác còn sót từ lần chạy trước (Render restart, crash...)
def _cleanup_stale_files():
    for fname in os.listdir(DOWNLOAD_DIR):
        fpath = os.path.join(DOWNLOAD_DIR, fname)
        try: os.remove(fpath)
        except Exception: pass

_cleanup_stale_files()

# Xóa file an toàn (không crash nếu file không tồn tại)
def safe_remove(path):
    if path and os.path.exists(path):
        try: os.remove(path)
        except Exception: pass

# session: { chat_id: { "eq": path | None, "mds": path | None, "pending": None } }
user_sessions = {}


# ─── Helper: tạo menu nút bấm chính ────────────────────
def main_menu():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(
        types.KeyboardButton("📊 Tạo Báo Cáo"),
        types.KeyboardButton("🔄 Làm Mới"),
        types.KeyboardButton("📋 Trạng Thái"),
        types.KeyboardButton("📜 Lịch Sử"),
        types.KeyboardButton("❓ Hướng Dẫn"),
    )
    return markup


# ─── Nhận dạng loại file: 3 tầng ────────────────────────
def detect_file_type(filename: str, local_path: str = None) -> str:
    """
    Nhận dạng file EQ hay MDS bằng cách ĐỌC NỘI DUNG bên trong.
    Tên file KHÔNG được dùng để phán đoán (quá dễ sai).
    
    Ưu tiên:
      1. Tên Sheet → chính xác nhất, ổn định
      2. Tiêu đề 5 dòng đầu → fallback nếu sheet tên chung chung
    Trả về: 'eq', 'mds', hoặc 'unknown'
    """
    eq_hints  = ['lý lịch', 'ly lich', 'equipment', 'ctbt', 'bảo trì thiết bị',
                 'lịch sử thiết bị', 'ly lich thiet bi']
    mds_hints = ['mds', 'nhật ký', 'nhat ky', 'maintenance', 'theo dõi ctbt']

    if not local_path:
        return 'unknown'

    try:
        import openpyxl, unicodedata

        def normalize(s):
            return unicodedata.normalize('NFC', str(s).lower())

        wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)

        # ── Bước 1: Tên Sheet ─────────────────────────────
        for s in wb.sheetnames:
            sn = normalize(s)
            if any(h in sn for h in eq_hints):  wb.close(); return 'eq'
            if any(h in sn for h in mds_hints): wb.close(); return 'mds'

        # ── Bước 2: Quét tiêu đề 5 dòng đầu ─────────────
        ws = wb.active
        header_text = ""
        for row in ws.iter_rows(max_row=5, values_only=True):
            for cell in row:
                if cell: header_text += normalize(cell) + " "
        wb.close()

        if any(h in header_text for h in eq_hints):  return 'eq'
        if any(h in header_text for h in mds_hints): return 'mds'

    except Exception as e:
        print(f"⚠️ Lỗi đọc file để nhận dạng: {e}")

    return 'unknown'


# ─── Helper: trạng thái session của 1 user ──────────────
def session_status_text(chat_id) -> str:
    s = user_sessions.get(chat_id, {})
    eq  = "✅ Đã có" if s.get("eq")  else "❌ Chưa nhận"
    mds = "✅ Đã có" if s.get("mds") else "❌ Chưa nhận"
    return (
        "📋 *Trạng thái phiên làm việc của bạn:*\n\n"
        f"📁 File Lý Lịch Thiết Bị:  {eq}\n"
        f"📁 File Nhật Ký MDS:         {mds}\n\n"
    ) + (
        "✔️ Bấm *📊 Tạo Báo Cáo* hoặc gửi file còn thiếu để tiến hành!"
        if s.get("eq") or s.get("mds")
        else "👉 Hãy gửi 2 file Excel để bắt đầu."
    )


# ─── /start ─────────────────────────────────────────────
@bot.message_handler(commands=['start'])
def send_welcome(message):
    user_sessions.setdefault(message.chat.id, {"eq": None, "mds": None})
    bot.send_message(
        message.chat.id,
        "👋 *Chào mừng đến với Bot Tổng Hợp Báo Cáo Bảo Trì BHS!*\n\n"
        "Chỉ cần gửi cho tôi *2 file Excel* nhật ký bảo trì,\n"
        "Bot sẽ tự phân loại và trả bạn file *Báo Cáo 4 Bảng* đẹp ngay lập tức! ⚡\n\n"
        "Dùng các *nút bên dưới* hoặc gửi file trực tiếp để bắt đầu.",
        parse_mode='Markdown',
        reply_markup=main_menu()
    )


# ─── Nút / lệnh văn bản ─────────────────────────────────
@bot.message_handler(func=lambda m: m.text in ["❓ Hướng Dẫn", "/help"])
def send_help(message):
    bot.send_message(
        message.chat.id,
        "📌 *Hướng dẫn sử dụng:*\n\n"
        "1️⃣  Gửi file *Lý Lịch Thiết Bị* (tên có chứa `eq` hoặc `equipment`)\n"
        "2️⃣  Gửi file *Nhật Ký MDS* (tên có chứa `mds` hoặc `maintenance`)\n"
        "3️⃣  Bấm *📊 Tạo Báo Cáo* – Bot tự xử lý & gửi trả file\n\n"
        "💡 *Lưu ý:*\n"
        "• Gửi 2 file theo *bất kỳ thứ tự nào* đều được\n"
        "• Bấm *🔄 Làm Mới* nếu muốn bắt đầu từ đầu\n"
        "• Bấm *📋 Trạng Thái* để xem đã nhận file chưa",
        parse_mode='Markdown',
        reply_markup=main_menu()
    )


@bot.message_handler(func=lambda m: m.text in ["📜 Lịch Sử", "/history"])
def send_history(message):
    bot.send_message(
        message.chat.id,
        history_manager.format_history(),
        parse_mode='Markdown',
        reply_markup=main_menu()
    )


@bot.message_handler(func=lambda m: m.text in ["📋 Trạng Thái", "/status"])
def send_status(message):
    bot.send_message(
        message.chat.id,
        session_status_text(message.chat.id),
        parse_mode='Markdown',
        reply_markup=main_menu()
    )


@bot.message_handler(func=lambda m: m.text in ["🔄 Làm Mới", "/reset"])
def reset_session(message):
    chat_id = message.chat.id
    s = user_sessions.get(chat_id, {})
    for key in ['eq', 'mds']:
        f = s.get(key)
        if f and os.path.exists(f):
            safe_remove(f)
    user_sessions[chat_id] = {"eq": None, "mds": None}
    bot.send_message(
        chat_id,
        "♻️ *Đã làm mới!* Sẵn sàng nhận 2 file mới từ bạn.",
        parse_mode='Markdown',
        reply_markup=main_menu()
    )


@bot.message_handler(func=lambda m: m.text in ["📊 Tạo Báo Cáo", "/generate"])
def trigger_generate(message):
    chat_id = message.chat.id
    s = user_sessions.get(chat_id, {"eq": None, "mds": None})

    if not s.get("eq") or not s.get("mds"):
        missing = []
        if not s.get("eq"):  missing.append("📁 File *Lý Lịch Thiết Bị*")
        if not s.get("mds"): missing.append("📁 File *Nhật Ký MDS*")
        bot.send_message(
            chat_id,
            "⚠️ Chưa đủ file để tạo báo cáo. Còn thiếu:\n\n" + "\n".join(missing),
            parse_mode='Markdown',
            reply_markup=main_menu()
        )
        return

    _run_report(chat_id, message)


# ─── Nhận file Excel ────────────────────────────────────
@bot.message_handler(content_types=['document'])
def handle_docs(message):
    chat_id = message.chat.id
    file_name = message.document.file_name

    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        bot.reply_to(message, "⚠️ Chỉ nhận file Excel (.xlsx / .xls)\nBấm *🔄 Làm Mới* nếu gửi nhầm.", parse_mode='Markdown')
        return

    user_sessions.setdefault(chat_id, {"eq": None, "mds": None, "pending": None})

    # Tải file xuống trước để có thể đọc nội dung
    try:
        file_info = bot.get_file(message.document.file_id)
        data = bot.download_file(file_info.file_path)
        tmp_path = os.path.join(DOWNLOAD_DIR, f"tmp_{chat_id}_{file_name}")
        with open(tmp_path, 'wb') as f:
            f.write(data)
    except Exception as e:
        bot.reply_to(message, f"❌ Lỗi tải file từ Telegram: `{e}`", parse_mode='Markdown')
        return

    # Nhận dạng 3 tầng: tên file → sheet names → header content
    file_type = detect_file_type(file_name, local_path=tmp_path)

    if file_type == 'unknown':
        # Vẫn không nhận ra → hỏi người dùng (file đã tải vào tmp, lưu path vào pending)
        user_sessions[chat_id]["pending"] = {
            "file_id": message.document.file_id,
            "file_name": file_name,
            "tmp_path": tmp_path       # Dùng lại, không tải lại lần nữa
        }
        markup = types.InlineKeyboardMarkup()
        markup.row(
            types.InlineKeyboardButton("🗂 File Lý Lịch TB", callback_data="settype:eq"),
            types.InlineKeyboardButton("🗂 File Nhật Ký MDS", callback_data="settype:mds")
        )
        bot.reply_to(
            message,
            f"🤔 Không nhận ra loại file `{file_name}`\nĐây là file nào?",
            parse_mode='Markdown',
            reply_markup=markup
        )
        return

    # Đã nhận dạng được → chuyển thẳng vào slot, không cần tải lại
    _store_file(chat_id, message, file_name, file_type, tmp_path)


# ─── Xử lý nút Inline (khi không nhận ra loại file) ────
@bot.callback_query_handler(func=lambda call: call.data.startswith("settype:"))
def callback_set_type(call):
    chat_id = call.message.chat.id
    file_type = call.data.split(":")[1]          # 'eq' hoặc 'mds' — an toàn, không bao giờ quá 64 bytes

    pending = user_sessions.get(chat_id, {}).get("pending")
    if not pending:
        bot.answer_callback_query(call.id, "⚠️ Phiên đã hết hạn. Gửi lại file nhé!")
        return

    label = 'Lý Lịch TB' if file_type == 'eq' else 'Nhật Ký MDS'
    bot.answer_callback_query(call.id, f"✅ Đã xác nhận: {label}")
    bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=None)

    user_sessions[chat_id]["pending"] = None
    _store_file(chat_id, call.message, pending["file_name"], file_type,
                pending.get("tmp_path"))          # Dùng lại file đã tải, không download lại


# ─── Lưu file đã tải vào session ───────────────────────
def _store_file(chat_id, message, file_name, file_type, tmp_path):
    """Đổi tên file tmp thành tên chuẩn và ghi vào session."""
    s = user_sessions.get(chat_id, {"eq": None, "mds": None, "pending": None})

    # Xóa file cũ cùng slot nếu có
    old = s.get(file_type)
    if old and os.path.exists(old):
        safe_remove(old)

    try:
        # Đổi tên file tạm thành tên chuẩn
        save_path = os.path.join(DOWNLOAD_DIR, f"user{chat_id}_{file_type}_{file_name}")
        if tmp_path and os.path.exists(tmp_path):
            os.rename(tmp_path, save_path)
        else:
            raise FileNotFoundError("File tạm không còn tồn tại.")

        user_sessions[chat_id][file_type] = save_path
        s = user_sessions[chat_id]

        label = "Lý Lịch Thiết Bị" if file_type == "eq" else "Nhật Ký MDS"
        has_both = s.get("eq") and s.get("mds")

        if has_both:
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("📊 Tạo Báo Cáo Ngay!", callback_data="generate_now"))
            bot.send_message(
                chat_id,
                f"✅ Đã nhận *{label}*: `{file_name}`\n\n"
                "🎯 *Đã đủ 2 file!* Bấm nút bên dưới để tạo báo cáo.",
                parse_mode='Markdown',
                reply_markup=markup
            )
        else:
            missing = "Lý Lịch Thiết Bị (EQ)" if not s.get("eq") else "Nhật Ký MDS"
            bot.send_message(
                chat_id,
                f"✅ Đã nhận *{label}*: `{file_name}`\n\n"
                f"📎 Còn thiếu file: *{missing}*",
                parse_mode='Markdown',
                reply_markup=main_menu()
            )

    except Exception as e:
        bot.send_message(chat_id, f"❌ Lỗi lưu file: `{e}`", parse_mode='Markdown')


# ─── Nút Inline "Tạo ngay" ──────────────────────────────
@bot.callback_query_handler(func=lambda call: call.data == "generate_now")
def callback_generate(call):
    bot.answer_callback_query(call.id, "Đang xử lý...")
    bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    _run_report(call.message.chat.id, call.message)


# ─── Hàm lõi: chạy generate_report ─────────────────────
def _run_report(chat_id, message):
    s = user_sessions.get(chat_id, {})
    file_eq  = s.get("eq")
    file_mds = s.get("mds")

    status_msg = bot.send_message(
        chat_id,
        "⚙️ *Đang xử lý...*\n\n"
        "▪️ Đọc dữ liệu...",
        parse_mode='Markdown'
    )

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = os.path.join(DOWNLOAD_DIR, f"Report_{chat_id}_{timestamp}.xlsx")

    try:
        bot.edit_message_text(
            "⚙️ *Đang xử lý...*\n\n"
            "▪️ Đọc dữ liệu... ✅\n"
            "▪️ Nhận dạng & phân loại băng chuyền...",
            chat_id, status_msg.message_id, parse_mode='Markdown'
        )

        success = generate_report.process_and_merge_excel(
            file_equipment=file_eq,
            file_mds=file_mds,
            output_excel=out_path,
            verbose=False
        )

        if success:
            bot.edit_message_text(
                "⚙️ *Đang xử lý...*\n\n"
                "▪️ Đọc dữ liệu... ✅\n"
                "▪️ Nhận dạng & phân loại... ✅\n"
                "▪️ Tô màu & xuất file... ✅",
                chat_id, status_msg.message_id, parse_mode='Markdown'
            )

            # Ghi lịch sử vào Google Sheets (Bỏ trống drive_link)
            uname = getattr(getattr(message, 'from_user', None), 'username', None)
            history_manager.add_entry(chat_id, uname, success, "")

            # Gửi kết quả cho user trực tiếp qua Telegram
            result_time = datetime.now().strftime('%d/%m/%Y %H:%M')
            caption = (
                f"🎉 *HOÀN TẤT!* _{result_time}_\n\n"
                "📋 BẢNG 1 \u2013 Băng Chuyền ĐI\n"
                "📋 BẢNG 2 \u2013 Băng Chuyền ĐẾN\n"
                "📋 BẢNG 3 \u2013 Check-In\n"
                "📋 BẢNG 4 \u2013 Sorter"
            )

            # Gửi đính kèm file trực tiếp
            with open(out_path, 'rb') as doc:
                bot.send_document(chat_id, doc, caption=caption, parse_mode='Markdown')

            bot.delete_message(chat_id, status_msg.message_id)
        else:
            bot.edit_message_text(
                "❌ Dữ liệu trong file trống hoặc không đúng định dạng.\nBấm *🔄 Làm Mới* để thử lại.",
                chat_id, status_msg.message_id, parse_mode='Markdown'
            )

    except Exception as e:
        bot.edit_message_text(
            f"❌ Lỗi khi xử lý:\n`{e}`\n\nBấm *🔄 Làm Mới* để thử lại.",
            chat_id, status_msg.message_id, parse_mode='Markdown'
        )

    finally:
        for key in ['eq', 'mds']:
            safe_remove(s.get(key))
        # Dọn cả file tmp pending nếu có (kẽ hở trước đây)
        pending = s.get('pending') or {}
        safe_remove(pending.get('tmp_path'))
        safe_remove(out_path)
        user_sessions[chat_id] = {"eq": None, "mds": None, "pending": None}


# ─── Bắt mọi tin nhắn text không khớp ──────────────────
@bot.message_handler(func=lambda m: True)
def catch_all(message):
    bot.send_message(
        message.chat.id,
        "👆 Dùng các nút bên dưới hoặc gửi file Excel trực tiếp nhé!",
        reply_markup=main_menu()
    )


# ─── KHỞI ĐỘNG ──────────────────────────────────────────
if __name__ == '__main__':
    keep_alive()
    print("🤖 Bot Telegram BHS đang chạy 24/7...")
    bot.infinity_polling(timeout=10, long_polling_timeout=5)
