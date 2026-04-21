import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import re
import warnings

FILE_EQUIPMENT = "/Users/phatdroid94/Desktop/xulyexcel/source_eq_history.xlsx"
FILE_MDS = "/Users/phatdroid94/Desktop/xulyexcel/source_mds_maintenance.xlsx"
OUTPUT_EXCEL = "/Users/phatdroid94/Desktop/xulyexcel/final_equipment_report.xlsx"

MAP_COLUMNS = {
    "Ngày/ tháng/ năm": "Ngày",
    "NỘI DUNG BẢO TRÌ/ SỬA CHỮA": "Nội dung bảo trì sửa chữa",
    "TÌNH TRẠNG HOẠT ĐỘNG TB trước bảo trì/sửa chữa": "Tình trạng của thiết bị trước khi sửa chữa",
    "VẬT TƯ THAY THẾ nếu có (tên, chủng loại, số lượng)": "Vật tư thay thế",
    "VẬT TƯ THAY THẾ nếu có\n(tên, chủng loại, số lượng)": "Vật tư thay thế",
    "SỐ PHIẾU CT (nếu có)": "Phiếu công tác",
    "PHIẾU KIỂM NGHIỆM KỸ THUẬT (nếu có)": "BB kiểm nghiệm",
    "TÌNH TRẠNG HOẠT ĐỘNG TB Sau bảo trì": "Tình trạng sau SC",
    "ĐƠN VỊ THỰC HIỆN (người thực hiện, ký, ghi rõ họ tên)": "Nhân sự thực hiện"
}

def format_excel_date(val):
    if pd.isna(val) or str(val).strip() == "": return ""
    try:
        if isinstance(val, (int, float)):
            dt = pd.Timedelta(days=float(val)) + pd.Timestamp('1899-12-30')
            return dt.strftime('%d/%m/%Y')
        
        # Tắt cảnh báo Parsing Dates rối mắt của Pandas
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            dt = pd.to_datetime(val, dayfirst=True)
            
        return dt.strftime('%d/%m/%Y')
    except Exception:
        return str(val)

def categorize_row(content, code):
    """Phân loại tự động dựa trên mô tả công việc và mã thiết bị"""
    text = str(content).lower() + " " + str(code).lower()
    
    # BẢNG 4: SORTER
    if any(k in text for k in ['sorter', 'ind', 'induction', 'chute', 'sm5', 'cctv', 'rơ le nhiệt']):
        return "BẢNG 4_SORTER"
    
    # BẢNG 3: CHECK IN
    if any(k in text for k in ['cân', 'loadcell', 'máy soi', 'check-in', 'quầy']):
        return "BẢNG 3_CHECK IN"
    if re.search(r'\b[a-l]\d+\b', str(code).lower()) and '-' not in str(code):
        return "BẢNG 3_CHECK IN"
        
    # BẢNG 2: BĂNG CHUYỀN ĐẾN
    if any(k in text for k in ['mcp a', 'băng chuyền đến']) or re.search(r'\ba\d+', str(code).lower()):
        return "BẢNG 2_BC ĐẾN"
        
    # BẢNG 1: BĂNG CHUYỀN ĐI
    if any(k in text for k in ['mcp', 'sm1', 'sm2', 'sm3', 'sm4', 'me']) or re.search(r'\bl\d+', str(code).lower()):
        return "BẢNG 1_BC ĐI"
        
    return "BẢNG 1_BC ĐI"

def read_legacy_sheet(file_path, sheet_name):
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=2)
        
        renames = {}
        for c in df.columns:
            for k, v in MAP_COLUMNS.items():
                if k.replace('\n', '') == str(c).replace('\n', ''):
                    renames[c] = v
        df.rename(columns=renames, inplace=True)
        
        target_columns = ["Ngày", "Nội dung bảo trì sửa chữa", "Tình trạng của thiết bị trước khi sửa chữa",
                          "Vật tư thay thế", "Phiếu công tác", "BB kiểm nghiệm", "Tình trạng sau SC", "Nhân sự thực hiện"]
        for col in target_columns:
            if col not in df.columns:
                df[col] = "" 
                
        df["Ngày"] = df["Ngày"].ffill()
        df = df.dropna(subset=["Nội dung bảo trì sửa chữa"], how='all')
        df["Ngày"] = df["Ngày"].apply(format_excel_date)
        df = df.fillna("")
        
        pattern = r'\b([A-Z]{1,4}[0-9]+(?:[A-Z]*-[A-Z0-9]+)*)\b'
        def extract_code(text):
            if pd.isna(text): return ""
            matches = re.findall(pattern, str(text))
            if not matches: return ""
            return ", ".join(list(dict.fromkeys(matches)))
        
        df["Ký hiệu băng chuyền"] = df["Nội dung bảo trì sửa chữa"].apply(extract_code)
        df['Sheet_Category'] = df.apply(lambda row: categorize_row(row["Nội dung bảo trì sửa chữa"], row["Ký hiệu băng chuyền"]), axis=1)
        df = df[target_columns + ['Sheet_Category']]
        return df
    except Exception as e:
        print(f"⚠️ [LỖI] Không thể đọc file: {e}")
        return pd.DataFrame()

def process_and_merge_excel(file_equipment=FILE_EQUIPMENT, file_mds=FILE_MDS, output_excel=OUTPUT_EXCEL, verbose=True) -> dict | bool:
    if verbose:
        print("=========================================================")
        print("🚀 BẮT ĐẦU TỔNG HỢP BÁO CÁO LÝ LỊCH THIẾT BỊ TỰ ĐỘNG 🚀")
        print("=========================================================")
        print("\n[Bước 1/4] 📖 Đang đọc dữ liệu từ các file Nhật ký...")
        
    df_eq = read_legacy_sheet(file_equipment, "Sổ theo dõi CTBT 2025")
    df_mds = read_legacy_sheet(file_mds, "Sổ theo dõi CTBT 2025")

    df_final = pd.concat([df_eq, df_mds], ignore_index=True)
    if df_final.empty:
        if verbose: print("❌ LỖI: Cả 2 file đều trống. Dừng chương trình.")
        return {}
        
    df_final = df_final.drop_duplicates()

    if verbose: print("[Bước 2/4] 🧠 Đang nhận diện Băng chuyền & Phân loại vào 4 Bảng...")
    sheets_dict = {
        "BẢNG 1_BC ĐI": df_final[df_final['Sheet_Category'] == 'BẢNG 1_BC ĐI'].drop(columns=['Sheet_Category']),
        "BẢNG 2_BC ĐẾN": df_final[df_final['Sheet_Category'] == 'BẢNG 2_BC ĐẾN'].drop(columns=['Sheet_Category']),
        "BẢNG 3_CHECK IN": df_final[df_final['Sheet_Category'] == 'BẢNG 3_CHECK IN'].drop(columns=['Sheet_Category']),
        "BẢNG 4_SORTER": df_final[df_final['Sheet_Category'] == 'BẢNG 4_SORTER'].drop(columns=['Sheet_Category'])
    }

    if verbose: print("[Bước 3/4] 💾 Đang kết xuất dữ liệu ra file Excel mới...")
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        for sheet_name, df_group in sheets_dict.items():
            df_group.to_excel(writer, index=False, sheet_name=sheet_name[:31]) 

    if verbose: print("[Bước 4/4] 🎨 Đang tô màu & Ép chuẩn UI Báo cáo...")
    wb = openpyxl.load_workbook(output_excel)
    
    colors = {
        "BẢNG 1_BC ĐI": "B4C6E7",
        "BẢNG 2_BC ĐẾN": "C6E0B4",
        "BẢNG 3_CHECK IN": "FCE4D6",
        "BẢNG 4_SORTER": "E2EFDA"
    }
    
    for ws_name in wb.sheetnames:
        sheet = wb[ws_name]
        fill_color = colors.get(ws_name, "D9D9D9")
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 40
        sheet.column_dimensions['D'].width = 40
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 30
        sheet.column_dimensions['H'].width = 25
        
    wb.save(output_excel)

    # Trả về số dòng từng sheet để Bot ghi lịch sử
    sheet_counts = {name.split("_", 1)[1]: len(df) for name, df in sheets_dict.items()}

    if verbose:
        print(f"\n🎉 HOÀN THẤT! Đã xuất 4 Bảng riêng biệt tại:\n{output_excel}")
        print("=========================================================\n")
    return sheet_counts

if __name__ == "__main__":
    process_and_merge_excel()